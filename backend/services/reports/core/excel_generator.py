"""
Excel Generator для DevAssist Pro
Генерация Excel отчетов по результатам анализа КП
"""
import asyncio
import logging
from typing import Dict, Any, Optional, List
from datetime import datetime
from pathlib import Path
import json
import uuid

# Заглушки для Excel библиотек (требуется установка)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.warning("OpenPyXL not installed. Excel generation will use mock implementation.")

from .report_generator import ReportGenerator

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Генератор Excel отчетов"""
    
    def __init__(self):
        self.report_generator = ReportGenerator()
        self.output_dir = Path("data/exports")
        self.output_dir.mkdir(exist_ok=True)
        
        # Настройка стилей
        if EXCEL_AVAILABLE:
            self._setup_styles()
    
    def _setup_styles(self):
        """Настройка стилей для Excel"""
        if not EXCEL_AVAILABLE:
            return
            
        # Стиль для заголовков
        self.header_font = Font(
            name='Arial',
            size=14,
            bold=True,
            color='FFFFFF'
        )
        
        self.header_fill = PatternFill(
            start_color='2E75D6',
            end_color='2E75D6',
            fill_type='solid'
        )
        
        # Стиль для подзаголовков
        self.subheader_font = Font(
            name='Arial',
            size=12,
            bold=True,
            color='1A1E3A'
        )
        
        self.subheader_fill = PatternFill(
            start_color='F4F7FC',
            end_color='F4F7FC',
            fill_type='solid'
        )
        
        # Стиль для обычного текста
        self.normal_font = Font(
            name='Arial',
            size=10
        )
        
        # Границы
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    async def generate_kp_analysis_report(
        self,
        analysis_id: int,
        include_charts: bool = True,
        include_raw_data: bool = False
    ) -> str:
        """
        Генерация Excel отчета по анализу КП
        
        Args:
            analysis_id: ID анализа
            include_charts: Включать диаграммы
            include_raw_data: Включать сырые данные
            
        Returns:
            Путь к сгенерированному Excel файлу
        """
        try:
            logger.info(f"Generating Excel report for analysis {analysis_id}")
            
            # Получение данных отчета
            report_data = await self.report_generator.generate_report(
                analysis_id=analysis_id,
                report_format="excel",
                include_charts=include_charts,
                include_raw_data=include_raw_data
            )
            
            # Генерация Excel файла
            filename = f"kp_analysis_{analysis_id}_{uuid.uuid4().hex[:8]}.xlsx"
            file_path = self.output_dir / filename
            
            if EXCEL_AVAILABLE:
                await self._generate_excel_file(report_data, file_path)
            else:
                await self._generate_mock_excel(report_data, file_path)
            
            logger.info(f"Excel report generated successfully: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise
    
    async def _generate_excel_file(self, report_data: Dict[str, Any], file_path: Path):
        """Генерация Excel файла с использованием OpenPyXL"""
        
        # Создание workbook
        wb = openpyxl.Workbook()
        
        # Удаление стандартного листа
        wb.remove(wb.active)
        
        # Создание листов
        await self._create_summary_sheet(wb, report_data)
        await self._create_analysis_sheet(wb, report_data)
        
        if "charts" in report_data:
            await self._create_charts_sheet(wb, report_data["charts"])
        
        if report_data["metadata"]["include_raw_data"]:
            await self._create_raw_data_sheet(wb, report_data)
        
        # Сохранение файла
        wb.save(file_path)
    
    async def _create_summary_sheet(self, wb, report_data: Dict[str, Any]):
        """Создание листа с резюме"""
        
        ws = wb.create_sheet("Резюме", 0)
        
        # Заголовок
        ws['A1'] = "Анализ коммерческого предложения"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Информация о генерации
        generated_at = datetime.fromisoformat(report_data["metadata"]["generated_at"])
        ws['A3'] = f"Сгенерировано: {generated_at.strftime('%d.%m.%Y %H:%M')}"
        ws['A3'].font = self.normal_font
        
        # Резюме анализа
        if "executive_summary" in report_data["sections"]:
            summary_data = report_data["sections"]["executive_summary"]
            
            ws['A5'] = "Резюме анализа"
            ws['A5'].font = self.subheader_font
            ws['A5'].fill = self.subheader_fill
            
            ws['A6'] = summary_data["content"]
            ws['A6'].font = self.normal_font
            ws['A6'].alignment = Alignment(wrap_text=True)
            
            # Ключевые выводы
            if "key_findings" in summary_data:
                ws['A8'] = "Ключевые выводы:"
                ws['A8'].font = self.subheader_font
                
                row = 9
                for finding in summary_data["key_findings"]:
                    ws[f'A{row}'] = f"• {finding}"
                    ws[f'A{row}'].font = self.normal_font
                    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                    row += 1
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    async def _create_analysis_sheet(self, wb, report_data: Dict[str, Any]):
        """Создание листа с детальным анализом"""
        
        ws = wb.create_sheet("Детальный анализ")
        
        row = 1
        
        # Заголовок
        ws[f'A{row}'] = "Детальный анализ"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 2
        
        # Информация о документе
        if "document_info" in report_data["sections"]:
            doc_info = report_data["sections"]["document_info"]["content"]
            
            ws[f'A{row}'] = "Информация о документе"
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = self.subheader_fill
            row += 1
            
            # Таблица с информацией
            headers = ["Параметр", "Значение"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
                ws.cell(row=row, column=col).font = self.subheader_font
                ws.cell(row=row, column=col).border = self.border
            row += 1
            
            doc_data = [
                ["Файл", doc_info.get("filename", "—")],
                ["Размер", f"{doc_info.get('file_size', 0) / 1024:.1f} KB"],
                ["Страниц", str(doc_info.get("pages", "—"))],
                ["Загружен", doc_info.get("uploaded_at", "—")[:10]]
            ]
            
            for data_row in doc_data:
                for col, value in enumerate(data_row, 1):
                    ws.cell(row=row, column=col, value=value)
                    ws.cell(row=row, column=col).font = self.normal_font
                    ws.cell(row=row, column=col).border = self.border
                row += 1
            
            row += 1
        
        # Результаты анализа
        if "analysis_results" in report_data["sections"]:
            results_data = report_data["sections"]["analysis_results"]
            
            ws[f'A{row}'] = "Результаты анализа"
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = self.subheader_fill
            row += 1
            
            # Техническое соответствие
            if "technical_requirements" in results_data:
                tech_req = results_data["technical_requirements"]
                
                ws[f'A{row}'] = "Техническое соответствие"
                ws[f'A{row}'].font = self.normal_font
                row += 1
                
                ws[f'A{row}'] = f"Оценка соответствия: {tech_req['compliance_score']}%"
                ws[f'A{row}'].font = self.normal_font
                row += 1
                
                if "missing_requirements" in tech_req:
                    ws[f'A{row}'] = "Отсутствующие требования:"
                    ws[f'A{row}'].font = self.normal_font
                    row += 1
                    
                    for req in tech_req["missing_requirements"]:
                        ws[f'A{row}'] = f"• {req}"
                        ws[f'A{row}'].font = self.normal_font
                        row += 1
                
                row += 1
            
            # Анализ стоимости
            if "cost_analysis" in results_data:
                cost_data = results_data["cost_analysis"]
                
                ws[f'A{row}'] = "Анализ стоимости"
                ws[f'A{row}'].font = self.normal_font
                row += 1
                
                ws[f'A{row}'] = f"Общая стоимость: {cost_data['total_cost']:,} руб."
                ws[f'A{row}'].font = self.normal_font
                row += 1
                
                ws[f'A{row}'] = f"Конкурентоспособность: {cost_data['competitiveness']}"
                ws[f'A{row}'].font = self.normal_font
                row += 2
                
                # Разбивка по статьям
                if "cost_breakdown" in cost_data:
                    ws[f'A{row}'] = "Разбивка затрат:"
                    ws[f'A{row}'].font = self.normal_font
                    row += 1
                    
                    for item, cost in cost_data["cost_breakdown"].items():
                        ws[f'A{row}'] = f"• {item}: {cost:,} руб."
                        ws[f'A{row}'].font = self.normal_font
                        row += 1
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    async def _create_charts_sheet(self, wb, charts_data: Dict[str, Any]):
        """Создание листа с диаграммами"""
        
        ws = wb.create_sheet("Диаграммы")
        
        # Заголовок
        ws['A1'] = "Диаграммы"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        row = 3
        
        for chart_name, chart_data in charts_data.items():
            if chart_data["type"] == "pie":
                row = await self._add_pie_chart_to_excel(ws, chart_data, row)
            elif chart_data["type"] == "gauge":
                row = await self._add_gauge_data_to_excel(ws, chart_data, row)
            
            row += 2
    
    async def _add_pie_chart_to_excel(self, ws, chart_data: Dict[str, Any], start_row: int) -> int:
        """Добавление круговой диаграммы в Excel"""
        
        # Заголовок диаграммы
        ws[f'A{start_row}'] = chart_data["title"]
        ws[f'A{start_row}'].font = self.subheader_font
        
        # Данные для диаграммы
        data_start_row = start_row + 1
        row = data_start_row
        
        for label, value in chart_data["data"].items():
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Создание диаграммы
        pie_chart = PieChart()
        pie_chart.title = chart_data["title"]
        
        # Данные
        data_ref = Reference(ws, min_col=2, min_row=data_start_row, max_row=row-1)
        labels_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=row-1)
        
        pie_chart.add_data(data_ref)
        pie_chart.set_categories(labels_ref)
        
        # Добавление диаграммы на лист
        ws.add_chart(pie_chart, f'D{start_row}')
        
        return row
    
    async def _add_gauge_data_to_excel(self, ws, chart_data: Dict[str, Any], start_row: int) -> int:
        """Добавление данных gauge диаграммы в Excel"""
        
        # Заголовок
        ws[f'A{start_row}'] = chart_data["title"]
        ws[f'A{start_row}'].font = self.subheader_font
        
        # Значения
        ws[f'A{start_row + 1}'] = "Значение"
        ws[f'B{start_row + 1}'] = chart_data["value"]
        
        ws[f'A{start_row + 2}'] = "Максимум"
        ws[f'B{start_row + 2}'] = chart_data["max_value"]
        
        ws[f'A{start_row + 3}'] = "Процент"
        ws[f'B{start_row + 3}'] = f"{(chart_data['value'] / chart_data['max_value']) * 100:.1f}%"
        
        return start_row + 4
    
    async def _create_raw_data_sheet(self, wb, report_data: Dict[str, Any]):
        """Создание листа с сырыми данными"""
        
        ws = wb.create_sheet("Сырые данные")
        
        # Заголовок
        ws['A1'] = "Сырые данные анализа"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        # Данные в формате JSON
        ws['A3'] = "JSON данные:"
        ws['A3'].font = self.subheader_font
        
        # Форматированный JSON
        json_data = json.dumps(report_data, ensure_ascii=False, indent=2)
        ws['A4'] = json_data
        ws['A4'].font = self.normal_font
        ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Настройка размеров
        ws.column_dimensions['A'].width = 100
        ws.row_dimensions[4].height = 500
    
    async def _generate_mock_excel(self, report_data: Dict[str, Any], file_path: Path):
        """Генерация мок Excel файла (когда OpenPyXL недоступен)"""
        
        # Простой CSV файл как заглушка
        content = []
        content.append("DevAssist Pro - Анализ коммерческого предложения")
        content.append("")
        
        generated_at = datetime.fromisoformat(report_data["metadata"]["generated_at"])
        content.append(f"Сгенерировано,{generated_at.strftime('%d.%m.%Y %H:%M')}")
        content.append("")
        
        # Резюме
        if "executive_summary" in report_data["sections"]:
            summary_data = report_data["sections"]["executive_summary"]
            content.append("Резюме анализа")
            content.append(f"\"{summary_data['content']}\"")
            content.append("")
            
            content.append("Ключевые выводы")
            for finding in summary_data.get("key_findings", []):
                content.append(f"\"{finding}\"")
        
        # Рекомендации
        if "recommendations" in report_data["sections"]:
            content.append("")
            content.append("Рекомендации")
            for rec in report_data["sections"]["recommendations"]["content"]:
                content.append(f"\"{rec}\"")
        
        # Замена расширения на .csv
        csv_path = file_path.with_suffix('.csv')
        
        # Запись в файл
        with open(csv_path, "w", encoding="utf-8") as f:
            f.write("\n".join(content))
        
        logger.info(f"Mock Excel (CSV) generated: {csv_path}")
        return str(csv_path)
    
    async def export_analysis_data(
        self,
        analysis_ids: List[int],
        format: str = "excel",
        include_details: bool = True
    ) -> str:
        """
        Экспорт данных анализа в Excel
        
        Args:
            analysis_ids: Список ID анализов
            format: Формат экспорта (excel, csv)
            include_details: Включать детальные данные
            
        Returns:
            Путь к файлу экспорта
        """
        try:
            logger.info(f"Exporting analysis data: {analysis_ids}")
            
            filename = f"analysis_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = self.output_dir / filename
            
            # Заглушка для экспорта
            if EXCEL_AVAILABLE:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Анализы"
                
                # Заголовки
                headers = ["ID", "Дата", "Статус", "Оценка", "Стоимость"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Данные (заглушка)
                for row, analysis_id in enumerate(analysis_ids, 2):
                    ws.cell(row=row, column=1, value=analysis_id)
                    ws.cell(row=row, column=2, value=datetime.now().strftime('%d.%m.%Y'))
                    ws.cell(row=row, column=3, value="Завершен")
                    ws.cell(row=row, column=4, value="85%")
                    ws.cell(row=row, column=5, value="2,500,000")
                
                wb.save(file_path)
            else:
                # Мок версия
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write("Mock Excel export file\n")
                    f.write(f"Analysis IDs: {', '.join(map(str, analysis_ids))}\n")
            
            logger.info(f"Analysis data exported: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Error exporting analysis data: {str(e)}")
            raise